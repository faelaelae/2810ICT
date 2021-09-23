# -*- coding: utf-8 -*-
"""
Created on Thu Sep 23 12:33:37 2021

@author: s2786134
"""

#Analysis 2

#For a user-selected period, produce a chart to show the distribution of cases in each offence code

import openpyxl
import datetime
import matplotlib.pyplot as plt
import numpy as np
from dateutil import parser
from collections import Counter

path = "testdata.xlsx"

def open_Workbook(path):
    wb = openpyxl.load_workbook(path)
    return wb

wb = open_Workbook(path)

sheet = wb['data']

startDate = datetime.datetime(2012,1,1)
endDate = datetime.datetime(2013,12,12)

def buildData(startDate, endDate, sheet):
    counter = 2
    data = []
    maxrows = sheet.max_row
    maxcolumn = sheet.max_column
    print(startDate, endDate)
    for row in sheet.iter_rows(min_row=1, max_col=maxcolumn,max_row=maxrows):
        counter += 1
        date = sheet.cell(row=counter,column=2).value
        if date is not None:
            #date = parser.parse(d)
            code = sheet.cell(row=counter,column=3).value
            print(code)
            if date >= startDate and date <= endDate:
                    data.append(code)
    print(data)
    return data
        
listData = buildData(startDate, endDate, sheet)
#def translateData(listData):
#    countedList = Counter(listData)
#    return countedList

data = Counter(listData)

def drawGraph(data):
    print(data.keys())
    print(data.values())
    index = np.arange(len(data.keys()))
    width = 0.2
    plt.bar(index, data.values(), width)
    plt.xticks(index + width * 0.5, data.keys())
    plt.show()

drawGraph(data)

    

